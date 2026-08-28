import tempfile
import unittest
from datetime import date
from decimal import Decimal
from pathlib import Path

from bank_statement_skill.export_xlsx import build_workbook
from bank_statement_skill.models import TransactionRecord


class XlsxOutputTests(unittest.TestCase):
    def test_build_workbook_has_traceable_sheets_and_formula_driven_months(self):
        records = [
            TransactionRecord(
                page=3,
                date=date(2024, 5, 22),
                txid="0163223376",
                description="应付商户延迟清算款",
                amount=Decimal("706.60"),
                balance=Decimal("6078.98"),
                raw_amount="706.60",
                raw_text="20240522 0163223376 应付商户延迟清算款 706.60 6,078.98",
                confidence="high",
                review_status="已提取",
            )
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "result.xlsx"
            build_workbook(records, {"source_file": "sample.pdf", "page_range": "1-3", "keywords": "应付商户延迟清算款"}, path)
            import openpyxl

            wb = openpyxl.load_workbook(path, data_only=False)
            self.assertEqual(wb.sheetnames, ["汇总", "按月统计", "匹配明细", "核验说明"])
            self.assertIn("COUNTA", wb["汇总"]["B7"].value)
            self.assertIn("COUNTIFS", wb["按月统计"]["B4"].value)
            self.assertEqual(wb["匹配明细"]["B2"].value, 3)
            self.assertEqual(Decimal(str(wb["匹配明细"]["F2"].value)), Decimal("706.60"))

    def test_build_workbook_keeps_empty_match_result_valid(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "empty.xlsx"
            build_workbook([], {"source_file": "sample.pdf", "page_range": "1-3", "keywords": "不存在"}, path)
            import openpyxl

            wb = openpyxl.load_workbook(path, data_only=False)
            self.assertIn("COUNTA", wb["汇总"]["B7"].value)
            self.assertEqual(wb["按月统计"]["B4"].value, "=0")
            self.assertEqual(wb["按月统计"]["C4"].value, "=0")

import os
import shutil
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path


class CliSmokeTests(unittest.TestCase):
    @unittest.skipUnless(shutil.which("tesseract"), "requires local tesseract")
    def test_cli_generates_xlsx_from_synthetic_scanned_pdf(self):
        try:
            import fitz
            from PIL import Image, ImageDraw, ImageFont
        except ImportError:
            self.skipTest("requires PyMuPDF and Pillow")
        with tempfile.TemporaryDirectory() as temp_dir:
            temp = Path(temp_dir)
            image_path = temp / "scan.png"
            pdf_path = temp / "scan.pdf"
            output_path = temp / "result.xlsx"
            image = Image.new("RGB", (1800, 500), "white")
            draw = ImageDraw.Draw(image)
            font = ImageFont.truetype("/System/Library/Fonts/Helvetica.ttc", 46) if Path("/System/Library/Fonts/Helvetica.ttc").exists() else ImageFont.load_default()
            draw.text((40, 70), "20240522 0163223376 delayed settlement payable 706.60 6078.98", fill="black", font=font)
            image.save(image_path)
            doc = fitz.open()
            page = doc.new_page(width=1800, height=500)
            page.insert_image(page.rect, filename=str(image_path))
            doc.save(pdf_path)
            doc.close()
            script = Path(__file__).parents[1] / "skills" / "bank-statement-scan-statistics" / "scripts" / "run_pipeline.py"
            env = os.environ.copy()
            env["PYTHONPATH"] = str(Path(__file__).parents[1] / "src")
            subprocess.run(
                [sys.executable, str(script), "--input", str(pdf_path), "--keyword", "delayed settlement", "--language", "eng", "--psm", "7", "--output", str(output_path)],
                check=True,
                env=env,
                capture_output=True,
                text=True,
            )
            self.assertTrue(output_path.exists())
            import openpyxl

            wb = openpyxl.load_workbook(output_path, data_only=False)
            self.assertIn("汇总", wb.sheetnames)
            self.assertGreaterEqual(wb["匹配明细"].max_row, 2)

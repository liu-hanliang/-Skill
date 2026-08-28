from dataclasses import dataclass
from pathlib import Path
from typing import List


@dataclass
class VerificationReport:
    path: Path
    sheets_ok: bool
    formulas_ok: bool
    detail_rows: int
    issues: List[str]

    @property
    def ok(self) -> bool:
        return self.sheets_ok and self.formulas_ok and not self.issues


def verify_workbook(path: Path) -> VerificationReport:
    try:
        import openpyxl  # type: ignore
    except ImportError as exc:
        raise RuntimeError("复核 XLSX 需要 openpyxl，请先安装项目依赖。") from exc
    path = Path(path).expanduser().resolve()
    wb = openpyxl.load_workbook(path, data_only=False, read_only=False)
    required = ["汇总", "按月统计", "匹配明细", "核验说明"]
    issues = []
    sheets_ok = all(name in wb.sheetnames for name in required)
    if not sheets_ok:
        issues.append(f"缺少工作表，当前为：{wb.sheetnames}")
    formulas_ok = True
    if "汇总" in wb.sheetnames:
        for cell in (wb["汇总"]["B7"], wb["汇总"]["E7"]):
            if not isinstance(cell.value, str) or not cell.value.startswith("="):
                formulas_ok = False
                issues.append(f"汇总公式缺失：{cell.coordinate}")
    if "按月统计" in wb.sheetnames:
        monthly = wb["按月统计"]
        if monthly.max_row < 4 or not str(monthly["B4"].value).startswith("="):
            formulas_ok = False
            issues.append("按月统计公式缺失：B4")
    detail_rows = max(wb["匹配明细"].max_row - 1, 0) if "匹配明细" in wb.sheetnames else 0
    return VerificationReport(path, sheets_ok, formulas_ok, detail_rows, issues)

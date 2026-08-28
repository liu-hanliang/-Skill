from collections import defaultdict
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from .models import TransactionRecord


def _require_openpyxl():
    try:
        import openpyxl  # type: ignore
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # type: ignore
        from openpyxl.worksheet.table import Table, TableStyleInfo  # type: ignore
        return openpyxl, Alignment, Border, Font, PatternFill, Side, Table, TableStyleInfo
    except ImportError as exc:
        raise RuntimeError("生成 XLSX 需要 openpyxl，请执行：python -m pip install -e '.[pdf]'") from exc


def build_workbook(records: Iterable[TransactionRecord], metadata: Dict[str, str], output_path: Path) -> Path:
    openpyxl, Alignment, Border, Font, PatternFill, Side, Table, TableStyleInfo = _require_openpyxl()
    records = list(records)
    output_path = Path(output_path).expanduser().resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    summary = wb.active
    summary.title = "汇总"
    monthly = wb.create_sheet("按月统计")
    detail = wb.create_sheet("匹配明细")
    notes = wb.create_sheet("核验说明")

    navy = "17365D"
    light_blue = "D9EAF7"
    light_green = "E2F0D9"
    light_yellow = "FFF2CC"
    light_gray = "F3F6FA"
    thin = Side(style="thin", color="D9E2F3")

    def style_title(ws, title, end_col):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
        cell = ws.cell(1, 1, title)
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(bold=True, color="FFFFFF", size=14)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28
        ws.sheet_view.showGridLines = False

    style_title(summary, "银行流水统计", 8)
    summary_rows = [
        (3, "统计对象", metadata.get("keywords", "")),
        (4, "统计范围", metadata.get("page_range", "")),
        (5, "来源文件", metadata.get("source_file", "")),
    ]
    for row, label, value in summary_rows:
        summary.cell(row, 1, label)
        summary.cell(row, 2, value)
        summary.cell(row, 1).fill = PatternFill("solid", fgColor=light_blue)
        summary.cell(row, 1).font = Font(bold=True, color=navy)
    summary.cell(7, 1, "匹配笔数")
    detail_end = max(len(records) + 1, 2)
    summary.cell(7, 2, f"=COUNTA('匹配明细'!$A$2:$A${detail_end})")
    summary.cell(7, 4, "总金额（元）")
    summary.cell(7, 5, f"=SUM('匹配明细'!$F$2:$F${detail_end})")
    summary.cell(7, 1).fill = PatternFill("solid", fgColor=light_green)
    summary.cell(7, 2).fill = PatternFill("solid", fgColor=light_green)
    summary.cell(7, 4).fill = PatternFill("solid", fgColor=light_yellow)
    summary.cell(7, 5).fill = PatternFill("solid", fgColor=light_yellow)
    summary.cell(7, 1).font = Font(bold=True, color=navy)
    summary.cell(7, 4).font = Font(bold=True, color=navy)
    summary.cell(7, 2).font = Font(bold=True, color="215E21", size=12)
    summary.cell(7, 5).font = Font(bold=True, color="7F6000", size=12)
    summary.cell(7, 5).number_format = "#,##0.00"
    summary.merge_cells("A9:H9")
    summary["A9"] = "统计结果用于证据整理和线索核对；提交诉讼、仲裁或审计前应逐页复核原始流水。"
    summary["A9"].fill = PatternFill("solid", fgColor=light_gray)
    summary["A9"].alignment = Alignment(wrap_text=True)
    summary.column_dimensions["A"].width = 18
    summary.column_dimensions["B"].width = 30
    summary.column_dimensions["D"].width = 18
    summary.column_dimensions["E"].width = 18

    style_title(monthly, "按月统计", 3)
    monthly.merge_cells("A2:C2")
    monthly["A2"] = "笔数和金额由匹配明细按交易日期区间公式汇总。"
    monthly["A2"].fill = PatternFill("solid", fgColor=light_gray)
    headers = ["月份", "笔数", "总金额（元）"]
    for col, value in enumerate(headers, 1):
        cell = monthly.cell(3, col, value)
        cell.fill = PatternFill("solid", fgColor=light_blue)
        cell.font = Font(bold=True, color=navy)
        cell.alignment = Alignment(horizontal="center")
    months = sorted({r.date.strftime("%Y-%m") for r in records if r.date})
    for offset, month in enumerate(months, 4):
        year, month_number = [int(x) for x in month.split("-")]
        if month_number == 12:
            next_year, next_month = year + 1, 1
        else:
            next_year, next_month = year, month_number + 1
        monthly.cell(offset, 1, month)
        monthly.cell(offset, 2, f"=COUNTIFS('匹配明细'!$C$2:$C${detail_end},\">=\"&DATE({year},{month_number},1),'匹配明细'!$C$2:$C${detail_end},\"<\"&DATE({next_year},{next_month},1))")
        monthly.cell(offset, 3, f"=SUMIFS('匹配明细'!$F$2:$F${detail_end},'匹配明细'!$C$2:$C${detail_end},\">=\"&DATE({year},{month_number},1),'匹配明细'!$C$2:$C${detail_end},\"<\"&DATE({next_year},{next_month},1))")
        monthly.cell(offset, 1).alignment = Alignment(horizontal="center")
        monthly.cell(offset, 3).number_format = "#,##0.00"
    total_row = 4 + len(months)
    monthly.cell(total_row, 1, "合计")
    if months:
        monthly.cell(total_row, 2, f"=SUM(B4:B{total_row-1})")
        monthly.cell(total_row, 3, f"=SUM(C4:C{total_row-1})")
    else:
        monthly.cell(total_row, 2, "=0")
        monthly.cell(total_row, 3, "=0")
    for col in range(1, 4):
        cell = monthly.cell(total_row, col)
        cell.fill = PatternFill("solid", fgColor=light_green)
        cell.font = Font(bold=True, color=navy)
        cell.border = Border(top=thin, bottom=thin)
    monthly.cell(total_row, 3).number_format = "#,##0.00"
    monthly.column_dimensions["A"].width = 16
    monthly.column_dimensions["B"].width = 12
    monthly.column_dimensions["C"].width = 18
    monthly.freeze_panes = "A4"
    if months:
        ref = f"A3:C{total_row-1}"
        tab = Table(displayName="MonthlySummary", ref=ref)
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        monthly.add_table(tab)

    detail_headers = ["序号", "PDF页码", "交易日期", "流水号（识别）", "匹配关键词", "交易金额（元）", "余额（元）", "OCR原始金额", "置信度", "核验状态", "原始OCR行", "备注"]
    for col, value in enumerate(detail_headers, 1):
        cell = detail.cell(1, col, value)
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for idx, record in enumerate(records, 1):
        values = [idx, record.page, record.date, record.txid, record.description, record.amount, record.balance, record.raw_amount, record.confidence, record.review_status, record.raw_text, record.note]
        for col, value in enumerate(values, 1):
            detail.cell(idx + 1, col, value)
        detail.cell(idx + 1, 3).number_format = "yyyy-mm-dd"
        detail.cell(idx + 1, 6).number_format = "#,##0.00"
        detail.cell(idx + 1, 7).number_format = "#,##0.00"
        if record.review_status != "已提取":
            for col in range(1, 13):
                detail.cell(idx + 1, col).fill = PatternFill("solid", fgColor="FFF2CC")
    widths = [8, 10, 14, 18, 24, 16, 16, 16, 10, 14, 60, 24]
    for col, width in enumerate(widths, 1):
        detail.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    detail.freeze_panes = "A2"
    detail.auto_filter.ref = f"A1:L{len(records)+1}"
    if records:
        tab = Table(displayName="MatchedRecords", ref=f"A1:L{len(records)+1}")
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        detail.add_table(tab)
    detail.sheet_view.showGridLines = False

    style_title(notes, "核验说明", 6)
    note_rows = [
        (3, "原始文件", metadata.get("source_file", "")),
        (4, "处理页数", metadata.get("page_range", "")),
        (5, "匹配关键词", metadata.get("keywords", "")),
        (6, "OCR方式", metadata.get("ocr_method", "本地 Tesseract")),
        (7, "人工复核", "低置信度、金额缺失、日期缺失、印章遮挡或版式未知的记录必须回看原始 PDF。"),
        (9, "隐私提示", "银行流水包含敏感信息。默认本地处理；不要把真实 PDF、OCR 文本或未经脱敏的 XLSX 提交到公开仓库。"),
    ]
    for row, label, value in note_rows:
        notes.cell(row, 1, label)
        notes.cell(row, 2, value)
        notes.cell(row, 1).fill = PatternFill("solid", fgColor=light_blue)
        notes.cell(row, 1).font = Font(bold=True, color=navy)
        notes.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        notes.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
    notes.column_dimensions["A"].width = 18
    for col in "BCDEF":
        notes.column_dimensions[col].width = 22
    notes.sheet_view.showGridLines = False

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except AttributeError:
        pass
    wb.save(output_path)
    return output_path
